import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import json
import logging

class ReportGenerator:
    """
    Generate comprehensive reports in various formats.
    Supports Excel, JSON, and HTML output formats.
    """
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        self.logger = logging.getLogger(__name__)
        os.makedirs(output_dir, exist_ok=True)
    
    def generate_excel_report(self, df: pd.DataFrame, summary_stats: dict, 
                            subscriber_analysis: dict, anomalies: list) -> str:
        """Generate comprehensive Excel report."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"balance_sync_report_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Sheet 1: Executive Summary
                self._create_summary_sheet(writer, summary_stats)
                
                # Sheet 2: Transaction Data
                df_clean = df.drop(['raw_log'], axis=1, errors='ignore')
                df_clean.to_excel(writer, sheet_name='Transaction Data', index=False)
                
                # Sheet 3: Overdraft Analysis
                if 'is_overdraft' in df.columns:
                    overdraft_df = df[df['is_overdraft']].copy()
                    overdraft_df.to_excel(writer, sheet_name='Overdrafts', index=False)
                
                # Sheet 4: Subscriber Analysis
                self._create_subscriber_sheet(writer, subscriber_analysis)
                
                # Sheet 5: Anomalies
                if anomalies:
                    anomaly_df = pd.DataFrame(anomalies)
                    anomaly_df.to_excel(writer, sheet_name='Anomalies', index=False)
                
                # Sheet 6: Recommendations
                self._create_recommendations_sheet(writer, summary_stats, subscriber_analysis)
            
            # Format the workbook
            self._format_excel_report(filepath)
            
            self.logger.info(f"Excel report generated: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error generating Excel report: {str(e)}")
            return ""
    
    def _create_summary_sheet(self, writer, summary_stats: dict):
        """Create executive summary sheet."""
        summary_data = []
        
        # Data overview
        data_overview = summary_stats.get('data_overview', {})
        summary_data.extend([
            ['EXECUTIVE SUMMARY', ''],
            ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['DATA OVERVIEW', ''],
            ['Total Transactions', data_overview.get('total_transactions', 0)],
            ['Date Range Start', data_overview.get('date_range', {}).get('start', 'N/A')],
            ['Date Range End', data_overview.get('date_range', {}).get('end', 'N/A')],
            ['Unique Subscribers', data_overview.get('unique_subscribers', 0)],
            ['Total Transaction Volume', f"${data_overview.get('total_transaction_volume', 0):,.2f}"],
            ['', '']
        ])
        
        # Transaction analysis
        trans_analysis = summary_stats.get('transaction_analysis', {})
        summary_data.extend([
            ['TRANSACTION ANALYSIS', ''],
            ['Average Transaction Amount', f"${trans_analysis.get('avg_transaction_amount', 0):,.2f}"],
            ['Median Transaction Amount', f"${trans_analysis.get('median_transaction_amount', 0):,.2f}"],
            ['Largest Transaction', f"${trans_analysis.get('largest_transaction', 0):,.2f}"],
            ['Smallest Transaction', f"${trans_analysis.get('smallest_transaction', 0):,.2f}"],
            ['', '']
        ])
        
        # Balance analysis
        balance_analysis = summary_stats.get('balance_analysis', {})
        summary_data.extend([
            ['BALANCE ANALYSIS', ''],
            ['Subscribers with Overdrafts', balance_analysis.get('subscribers_with_overdrafts', 0)],
            ['Total Overdraft Instances', balance_analysis.get('total_overdraft_instances', 0)],
            ['Average Running Balance', f"${balance_analysis.get('avg_running_balance', 0):,.2f}"],
            ['', '']
        ])
        
        # Operational metrics
        ops_metrics = summary_stats.get('operational_metrics', {})
        summary_data.extend([
            ['OPERATIONAL METRICS', ''],
            ['Success Rate', f"{ops_metrics.get('success_rate', 0):.1f}%"],
            ['Average Processing Time', f"{ops_metrics.get('avg_processing_time', 0):.1f} ms"],
        ])
        
        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
    
    def _create_subscriber_sheet(self, writer, subscriber_analysis: dict):
        """Create subscriber analysis sheet."""
        subscriber_data = []
        
        for subscriber_id, analysis in subscriber_analysis.items():
            subscriber_data.append([
                subscriber_id,
                analysis.get('total_transactions', 0),
                f"${analysis.get('total_volume', 0):,.2f}",
                f"${analysis.get('avg_transaction_size', 0):,.2f}",
                analysis.get('transaction_frequency', 'unknown'),
                analysis.get('preferred_transaction_type', 'unknown'),
                f"{analysis.get('risk_score', 0):.1f}",
                analysis.get('balance_stability', 'unknown')
            ])
        
        columns = [
            'Subscriber ID', 'Total Transactions', 'Total Volume', 
            'Avg Transaction Size', 'Frequency', 'Preferred Type',
            'Risk Score', 'Balance Stability'
        ]
        
        subscriber_df = pd.DataFrame(subscriber_data, columns=columns)
        subscriber_df.to_excel(writer, sheet_name='Subscriber Analysis', index=False)
    
    def _create_recommendations_sheet(self, writer, summary_stats: dict, subscriber_analysis: dict):
        """Create recommendations sheet."""
        recommendations = []
        
        # Overdraft recommendations
        overdraft_count = summary_stats.get('balance_analysis', {}).get('total_overdraft_instances', 0)
        if overdraft_count > 0:
            recommendations.append([
                'High Priority',
                'Overdraft Management',
                f'Found {overdraft_count} overdraft instances',
                'Implement real-time balance monitoring and automated alerts'
            ])
        
        # High-risk subscriber recommendations
        high_risk_subs = [
            sub_id for sub_id, analysis in subscriber_analysis.items()
            if analysis.get('risk_score', 0) > 70
        ]
        
        if high_risk_subs:
            recommendations.append([
                'High Priority',
                'Risk Management',
                f'{len(high_risk_subs)} high-risk subscribers identified',
                'Enhanced monitoring and credit limit reviews recommended'
            ])
        
        # System performance recommendations
        success_rate = summary_stats.get('operational_metrics', {}).get('success_rate', 100)
        if success_rate < 95:
            recommendations.append([
                'Medium Priority',
                'System Reliability',
                f'Transaction success rate: {success_rate:.1f}%',
                'Investigate and resolve system reliability issues'
            ])
        
        # General recommendations
        recommendations.extend([
            ['Medium Priority', 'Process Improvement', 'Manual reconciliation burden', 'Implement automated daily reconciliation'],
            ['Low Priority', 'Analytics Enhancement', 'Limited trend analysis', 'Deploy predictive analytics for fraud detection'],
            ['Low Priority', 'Reporting', 'Manual report generation', 'Automate weekly executive dashboards']
        ])
        
        rec_df = pd.DataFrame(recommendations, columns=['Priority', 'Category', 'Issue', 'Recommendation'])
        rec_df.to_excel(writer, sheet_name='Recommendations', index=False)
    
    def _format_excel_report(self, filepath: str):
        """Apply formatting to Excel report."""
        try:
            workbook = openpyxl.load_workbook(filepath)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_font = Font(bold=True, size=14)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Format headers
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Format title cells (cells with all caps text)
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.isupper():
                            cell.font = title_font
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(filepath)
            
        except Exception as e:
            self.logger.warning(f"Error formatting Excel file: {str(e)}")
    
    def generate_json_report(self, df: pd.DataFrame, summary_stats: dict, 
                           subscriber_analysis: dict, anomalies: list) -> str:
        """Generate JSON report for API consumption."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"balance_sync_report_{timestamp}.json"
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            # Convert DataFrame to records (excluding raw logs for size)
            df_clean = df.drop(['raw_log'], axis=1, errors='ignore')
            transactions = df_clean.to_dict('records')
            
            # Convert datetime objects to strings
            for transaction in transactions:
                for key, value in transaction.items():
                    if isinstance(value, datetime):
                        transaction[key] = value.isoformat()
                    elif pd.isna(value):
                        transaction[key] = None
            
            report_data = {
                'report_metadata': {
                    'generated_at': datetime.now().isoformat(),
                    'report_type': 'balance_sync_analysis',
                    'data_period': {
                        'start': summary_stats.get('data_overview', {}).get('date_range', {}).get('start'),
                        'end': summary_stats.get('data_overview', {}).get('date_range', {}).get('end')
                    }
                },
                'summary_statistics': summary_stats,
                'subscriber_analysis': subscriber_analysis,
                'anomalies': anomalies,
                'transaction_sample': transactions[:100],  # First 100 transactions
                'total_transactions': len(transactions)
            }
            
            with open(filepath, 'w') as f:
                json.dump(report_data, f, indent=2, default=str)
            
            self.logger.info(f"JSON report generated: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error generating JSON report: {str(e)}")
            return ""
    
    def generate_html_summary(self, summary_stats: dict) -> str:
        """Generate HTML summary for web display."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"summary_{timestamp}.html"
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            data_overview = summary_stats.get('data_overview', {})
            balance_analysis = summary_stats.get('balance_analysis', {})
            
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Balance Sync Analysis Summary</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    .header {{ background-color: #366092; color: white; padding: 20px; text-align: center; }}
                    .metric {{ margin: 10px 0; padding: 10px; border-left: 4px solid #366092; }}
                    .alert {{ background-color: #ffebee; border-left: 4px solid #f44336; }}
                    .success {{ background-color: #e8f5e8; border-left: 4px solid #4caf50; }}
                    .table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                    .table th, .table td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    .table th {{ background-color: #f2f2f2; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>Calo Balance Sync Analysis</h1>
                    <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
                
                <h2>Key Metrics</h2>
                <div class="metric">
                    <strong>Total Transactions:</strong> {data_overview.get('total_transactions', 0):,}
                </div>
                <div class="metric">
                    <strong>Unique Subscribers:</strong> {data_overview.get('unique_subscribers', 0):,}
                </div>
                <div class="metric">
                    <strong>Total Volume:</strong> ${data_overview.get('total_transaction_volume', 0):,.2f}
                </div>
                
                <h2>Balance Health</h2>
                <div class="metric {'alert' if balance_analysis.get('subscribers_with_overdrafts', 0) > 0 else 'success'}">
                    <strong>Subscribers with Overdrafts:</strong> {balance_analysis.get('subscribers_with_overdrafts', 0)}
                </div>
                <div class="metric {'alert' if balance_analysis.get('total_overdraft_instances', 0) > 0 else 'success'}">
                    <strong>Total Overdraft Instances:</strong> {balance_analysis.get('total_overdraft_instances', 0)}
                </div>
                
                <h2>System Performance</h2>
                <div class="metric">
                    <strong>Success Rate:</strong> {summary_stats.get('operational_metrics', {}).get('success_rate', 0):.1f}%
                </div>
                <div class="metric">
                    <strong>Average Processing Time:</strong> {summary_stats.get('operational_metrics', {}).get('avg_processing_time', 0):.1f} ms
                </div>
            </body>
            </html>
            """
            
            with open(filepath, 'w') as f:
                f.write(html_content)
            
            self.logger.info(f"HTML summary generated: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error generating HTML summary: {str(e)}")
            return ""
